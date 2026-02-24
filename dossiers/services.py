"""
RiskGuard 360 - Services
=========================
Couche service reliant le moteur de scoring aux modèles Django.
Inclut : scoring, audit, PDF pro, export, notifications, tâches async.
"""

import threading
import os
import sys
from decimal import Decimal
from io import BytesIO
from django.utils import timezone
from django.conf import settings
from django.core.mail import send_mail

sys.path.insert(0, str(settings.BASE_DIR))

from scoring_engine.scoring import (
    ProfilClient, calculer_score_risque, ResultatScoring as ScoringResult,
    simuler_pret
)
from .models import (
    DossierPret, ResultatScoring, AuditLog, Client, Notification
)


def get_client_ip(request):
    """Récupère l'adresse IP du client."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0]
    return request.META.get('REMOTE_ADDR')


def creer_audit_log(utilisateur, action, modele, objet_id, description,
                    donnees_avant=None, donnees_apres=None, raison="",
                    adresse_ip=None):
    """Crée une entrée dans le journal d'audit."""
    AuditLog.objects.create(
        utilisateur=utilisateur,
        action=action,
        modele=modele,
        objet_id=str(objet_id),
        description=description,
        donnees_avant=donnees_avant,
        donnees_apres=donnees_apres,
        raison=raison,
        adresse_ip=adresse_ip,
    )


def creer_notification(destinataire, titre, message, type_notif='info',
                       lien='', envoyer_email=False):
    """Crée une notification et optionnellement envoie un email."""
    notif = Notification.objects.create(
        destinataire=destinataire,
        type_notif=type_notif,
        titre=titre,
        message=message,
        lien=lien,
    )

    if envoyer_email and destinataire.email:
        try:
            send_mail(
                subject=f"[RiskGuard 360] {titre}",
                message=message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[destinataire.email],
                fail_silently=True,
            )
            notif.envoyee_email = True
            notif.save()
        except Exception:
            pass

    return notif


def calculer_score_dossier(dossier: DossierPret, utilisateur=None, adresse_ip=None):
    """
    Calcule le score de risque d'un dossier.
    Inclut score fraude + explication IA.
    """
    client = dossier.client

    profil = ProfilClient(
        nom=client.nom,
        prenom=client.prenom or "",
        type_client=client.type_client,
        profession=client.profession,
        revenu_mensuel=float(client.revenu_mensuel),
        charges_mensuelles=float(client.charges_mensuelles),
        dettes_existantes=float(client.dettes_existantes),
        anciennete_emploi=float(client.anciennete_emploi),
        incidents_paiement_12m=client.incidents_paiement,
        montant_demande=float(dossier.montant_demande),
        duree_pret_mois=dossier.duree_mois,
        age=client.age,
        apport_personnel=float(dossier.apport_personnel),
        numero_cni=getattr(client, 'numero_cni', ''),
    )

    resultat = calculer_score_risque(profil)

    donnees_avant = None
    if dossier.score_risque is not None:
        donnees_avant = {
            'score_risque': str(dossier.score_risque),
            'niveau_risque': dossier.niveau_risque,
        }

    # Mettre à jour le dossier
    dossier.score_risque = Decimal(str(resultat.score_global))
    dossier.score_fraude = Decimal(str(resultat.score_fraude))
    dossier.niveau_risque = resultat.niveau_risque.value
    dossier.recommandation = resultat.recommandation
    dossier.explication_score = resultat.explication
    dossier.alerte_fraude = resultat.alerte_fraude
    dossier.details_scoring = resultat.to_dict()
    dossier.save()

    # Créer l'enregistrement
    resultat_db = ResultatScoring.objects.create(
        dossier=dossier,
        score_global=Decimal(str(resultat.score_global)),
        score_endettement=Decimal(str(resultat.score_endettement)),
        score_historique=Decimal(str(resultat.score_historique)),
        score_stabilite=Decimal(str(resultat.score_stabilite)),
        score_coherence=Decimal(str(resultat.score_coherence)),
        score_fraude=Decimal(str(resultat.score_fraude)),
        ratio_endettement=Decimal(str(resultat.ratio_endettement)),
        niveau_risque=resultat.niveau_risque.value,
        recommandation=resultat.recommandation,
        explication=resultat.explication,
        alerte_fraude=resultat.alerte_fraude,
        alertes=resultat.alertes,
        details=resultat.details,
        calcule_par=utilisateur,
    )

    # Audit
    creer_audit_log(
        utilisateur=utilisateur,
        action='calcul_score',
        modele='DossierPret',
        objet_id=dossier.id,
        description=f"Score calculé pour {dossier.reference}: "
                    f"{resultat.score_global}/100 ({resultat.niveau_risque.value}) "
                    f"| Fraude: {resultat.score_fraude}/100",
        donnees_avant=donnees_avant,
        donnees_apres={
            'score_risque': str(resultat.score_global),
            'score_fraude': str(resultat.score_fraude),
            'niveau_risque': resultat.niveau_risque.value,
        },
        adresse_ip=adresse_ip,
    )

    # Notification si fraude
    if resultat.alerte_fraude and utilisateur:
        creer_notification(
            destinataire=utilisateur,
            titre=f"⚠️ Alerte Fraude - {dossier.reference}",
            message=f"Le moteur de scoring a détecté une fraude potentielle "
                    f"sur le dossier {dossier.reference}.\n"
                    f"Score fraude: {resultat.score_fraude}/100\n"
                    f"Client: {client}",
            type_notif='danger',
            lien=f'/dossiers/{dossier.pk}/',
            envoyer_email=True,
        )

    return resultat_db


def calculer_score_async(dossier_id, utilisateur_id=None):
    """Lance le calcul du score dans un thread séparé."""
    def _calcul():
        from django.contrib.auth import get_user_model
        User = get_user_model()
        try:
            dossier = DossierPret.objects.get(id=dossier_id)
            utilisateur = User.objects.get(id=utilisateur_id) if utilisateur_id else None
            calculer_score_dossier(dossier, utilisateur)
        except Exception as e:
            print(f"Erreur calcul asynchrone : {e}")

    thread = threading.Thread(target=_calcul)
    thread.daemon = True
    thread.start()
    return thread


def changer_etat_dossier(dossier: DossierPret, nouvel_etat: str,
                         utilisateur=None, motif="", adresse_ip=None):
    """Change l'état d'un dossier avec validation des transitions."""
    if nouvel_etat not in dossier.transitions_possibles:
        return False

    ancien_etat = dossier.etat
    dossier.etat = nouvel_etat
    dossier.motif_decision = motif

    if nouvel_etat == 'en_analyse':
        dossier.date_analyse = timezone.now()
    elif nouvel_etat in ('valide', 'refuse', 'alerte_fraude'):
        dossier.date_decision = timezone.now()

    dossier.save()

    creer_audit_log(
        utilisateur=utilisateur,
        action='changement_etat',
        modele='DossierPret',
        objet_id=dossier.id,
        description=f"Dossier {dossier.reference}: {ancien_etat} → {nouvel_etat}",
        donnees_avant={'etat': ancien_etat},
        donnees_apres={'etat': nouvel_etat},
        raison=motif,
        adresse_ip=adresse_ip,
    )

    # Notification au conseiller
    if dossier.conseiller:
        etat_display = dict(DossierPret.ETAT_CHOICES).get(nouvel_etat, nouvel_etat)
        type_notif = 'success' if nouvel_etat == 'valide' else \
                     'danger' if nouvel_etat in ('refuse', 'alerte_fraude') else 'info'
        creer_notification(
            destinataire=dossier.conseiller,
            titre=f"Dossier {dossier.reference} → {etat_display}",
            message=f"Le dossier {dossier.reference} ({dossier.client}) "
                    f"est passé à l'état : {etat_display}.\nMotif : {motif}",
            type_notif=type_notif,
            lien=f'/dossiers/{dossier.pk}/',
            envoyer_email=True,
        )

    return True


def generer_rapport_pdf(dossier: DossierPret):
    """
    Génère un rapport de risque professionnel en PDF avec QR Code.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm, mm
    from reportlab.lib.colors import HexColor, black, white, Color
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, KeepTogether, Image
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    import qrcode

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=2 * cm, bottomMargin=2 * cm,
        leftMargin=2 * cm, rightMargin=2 * cm
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='TitreRapport', parent=styles['Title'],
        fontSize=24, textColor=HexColor('#0d1b2a'),
        spaceAfter=5, alignment=TA_CENTER, fontName='Helvetica-Bold',
    ))
    styles.add(ParagraphStyle(
        name='Subtitle', parent=styles['Normal'],
        fontSize=11, textColor=HexColor('#415a77'),
        spaceAfter=15, alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        name='SousTitre', parent=styles['Heading2'],
        fontSize=13, textColor=HexColor('#1b263b'),
        spaceBefore=15, spaceAfter=8, fontName='Helvetica-Bold',
    ))
    styles.add(ParagraphStyle(
        name='Info', parent=styles['Normal'], fontSize=10, spaceAfter=4,
    ))
    styles.add(ParagraphStyle(
        name='Alerte', parent=styles['Normal'],
        fontSize=10, textColor=HexColor('#c62828'), spaceAfter=4,
    ))

    elements = []

    # ─── QR CODE ───
    qr_data = f"RiskGuard360|{dossier.reference}|Score:{dossier.score_risque}|{timezone.now().isoformat()}"
    qr = qrcode.QRCode(version=1, box_size=4, border=2)
    qr.add_data(qr_data)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="#0d1b2a", back_color="white")
    qr_buffer = BytesIO()
    qr_img.save(qr_buffer, format='PNG')
    qr_buffer.seek(0)

    # ─── EN-TÊTE avec QR ───
    header_data = [[
        Paragraph(
            '<font size="24" color="#0d1b2a"><b>RISKGUARD 360</b></font><br/>'
            '<font size="10" color="#415a77">Système d\'Analyse et de Scoring Risque Client</font>',
            styles['Info']
        ),
        Image(qr_buffer, width=2.5 * cm, height=2.5 * cm)
    ]]
    header_table = Table(header_data, colWidths=[12 * cm, 3 * cm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 3))

    # Ligne de séparation
    elements.append(HRFlowable(width="100%", thickness=3, color=HexColor('#0d1b2a')))
    elements.append(Spacer(1, 5))

    # Watermark info
    elements.append(Paragraph(
        f'<font size="8" color="#778da9">Réf: {dossier.reference} | '
        f'Généré le {timezone.now().strftime("%d/%m/%Y à %H:%M")} | '
        f'Document confidentiel</font>',
        ParagraphStyle('WM', alignment=TA_CENTER)
    ))
    elements.append(Spacer(1, 15))

    # ─── SCORE EN GROS ───
    if dossier.score_risque is not None:
        score = float(dossier.score_risque)
        if score >= 65:
            couleur = '#2e7d32'
        elif score >= 50:
            couleur = '#f57f17'
        elif score >= 35:
            couleur = '#e65100'
        else:
            couleur = '#b71c1c'

        score_fraude = float(dossier.score_fraude or 0)
        score_display = [
            [
                Paragraph(f'<font size="36" color="{couleur}"><b>{dossier.score_risque}/100</b></font>', styles['Info']),
                Paragraph(f'<font size="14"><b>Risque: {dossier.niveau_risque}</b></font><br/>'
                          f'<font size="10" color="#778da9">Score fraude: {score_fraude}/100</font>', styles['Info']),
            ]
        ]
        score_table = Table(score_display, colWidths=[7 * cm, 8 * cm])
        score_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, -1), HexColor('#f8f9fa')),
            ('ROUNDEDCORNERS', [8, 8, 8, 8]),
            ('TOPPADDING', (0, 0), (-1, -1), 15),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
            ('LEFTPADDING', (0, 0), (-1, -1), 20),
        ]))
        elements.append(score_table)
        elements.append(Spacer(1, 15))

    # ─── INFOS DOSSIER ───
    elements.append(Paragraph("📋 Informations du Dossier", styles['SousTitre']))
    infos = [
        ['Référence', dossier.reference, 'État', dossier.get_etat_display()],
        ['Objet du prêt', dossier.get_objet_pret_display(), 'Date soumission', dossier.date_soumission.strftime('%d/%m/%Y')],
        ['Montant demandé', f"{dossier.montant_demande:,.0f} FCFA", 'Durée', f"{dossier.duree_mois} mois"],
        ['Apport personnel', f"{dossier.apport_personnel:,.0f} FCFA", 'Conseiller', str(dossier.conseiller or '-')],
    ]
    t = Table(infos, colWidths=[4 * cm, 4.5 * cm, 3.5 * cm, 4 * cm])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (0, -1), HexColor('#415a77')),
        ('TEXTCOLOR', (2, 0), (2, -1), HexColor('#415a77')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.3, HexColor('#e0e0e0')),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 12))

    # ─── PROFIL CLIENT ───
    client = dossier.client
    elements.append(Paragraph("👤 Profil Client", styles['SousTitre']))
    client_info = [
        ['Client', str(client), 'Type', client.get_type_client_display()],
        ['Profession', client.get_profession_display(), 'Ancienneté', f"{client.anciennete_emploi} ans"],
        ['Revenu mensuel', f"{client.revenu_mensuel:,.0f} FCFA", 'Charges', f"{client.charges_mensuelles:,.0f} FCFA"],
        ['Dettes', f"{client.dettes_existantes:,.0f} FCFA", 'Incidents (12m)', str(client.incidents_paiement)],
    ]
    t2 = Table(client_info, colWidths=[4 * cm, 4.5 * cm, 3.5 * cm, 4 * cm])
    t2.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (0, -1), HexColor('#415a77')),
        ('TEXTCOLOR', (2, 0), (2, -1), HexColor('#415a77')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.3, HexColor('#e0e0e0')),
    ]))
    elements.append(t2)
    elements.append(Spacer(1, 12))

    # ─── DÉTAILS SCORING ───
    if dossier.score_risque is not None:
        elements.append(Paragraph("📊 Détails du Scoring", styles['SousTitre']))
        details = dossier.details_scoring or {}

        scores_detail = [
            ['Critère', 'Score', 'Poids', 'Contribution'],
            ['Ratio d\'endettement', f"{details.get('score_endettement', '-')}/100", '35%',
             f"{float(details.get('score_endettement', 0)) * 0.35:.1f}"],
            ['Historique paiement', f"{details.get('score_historique', '-')}/100", '30%',
             f"{float(details.get('score_historique', 0)) * 0.30:.1f}"],
            ['Stabilité professionnelle', f"{details.get('score_stabilite', '-')}/100", '20%',
             f"{float(details.get('score_stabilite', 0)) * 0.20:.1f}"],
            ['Cohérence montant', f"{details.get('score_coherence', '-')}/100", '15%',
             f"{float(details.get('score_coherence', 0)) * 0.15:.1f}"],
        ]
        t3 = Table(scores_detail, colWidths=[5.5 * cm, 3 * cm, 2.5 * cm, 3 * cm])
        t3.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#0d1b2a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#bdbdbd')),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, HexColor('#f5f5f5')]),
        ]))
        elements.append(t3)
        elements.append(Spacer(1, 10))

        # Ratio + Mensualité
        ratio = details.get('ratio_endettement', 0)
        mensualite = details.get('mensualite_estimee', 0)
        elements.append(Paragraph(
            f"<b>Ratio d'endettement:</b> {ratio * 100 if ratio else 0:.1f}% | "
            f"<b>Mensualité estimée:</b> {mensualite:,.0f} FCFA | "
            f"<b>Taux annuel:</b> 15%",
            styles['Info']
        ))
        elements.append(Spacer(1, 10))

        # ─── EXPLICATION IA ───
        explication = details.get('explication', '') or dossier.explication_score
        if explication:
            elements.append(Paragraph("🧠 Analyse Intelligente (Explainable AI)", styles['SousTitre']))
            for ligne in explication.split('\n'):
                elements.append(Paragraph(ligne, styles['Info']))
            elements.append(Spacer(1, 10))

        # Alertes
        alertes = details.get('alertes', [])
        if alertes:
            elements.append(Paragraph("⚠️ Alertes", styles['SousTitre']))
            for alerte in alertes:
                elements.append(Paragraph(f"• {alerte}", styles['Alerte']))
            elements.append(Spacer(1, 10))

        # Recommandation
        elements.append(Paragraph("✅ Recommandation", styles['SousTitre']))
        rec_color = couleur
        elements.append(Paragraph(
            f'<font color="{rec_color}" size="11"><b>{dossier.recommandation}</b></font>',
            styles['Info']
        ))

        if dossier.alerte_fraude:
            elements.append(Spacer(1, 10))
            elements.append(Paragraph(
                '<font color="#b71c1c" size="14"><b>🚨 ALERTE FRAUDE DÉTECTÉE 🚨</b></font>',
                ParagraphStyle('FraudeAlert', alignment=TA_CENTER)
            ))

    # ─── PIED DE PAGE ───
    elements.append(Spacer(1, 30))
    elements.append(HRFlowable(width="100%", thickness=1, color=HexColor('#bdbdbd')))
    elements.append(Spacer(1, 5))
    elements.append(Paragraph(
        f'<font size="7" color="#778da9">'
        f'RiskGuard 360 — Système d\'Analyse et de Scoring Risque Client — Document confidentiel<br/>'
        f'Généré le {timezone.now().strftime("%d/%m/%Y à %H:%M:%S")} — '
        f'Vérifiez l\'authenticité via le QR code ci-dessus</font>',
        ParagraphStyle('Footer', alignment=TA_CENTER)
    ))

    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def exporter_dossiers_excel(queryset):
    """
    Exporte les dossiers en fichier Excel.
    Returns: bytes du fichier Excel
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dossiers de Prêt"

    # Style en-tête
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # En-têtes
    headers = [
        'Référence', 'Client', 'Type Client', 'Montant (FCFA)',
        'Durée (mois)', 'Objet', 'État', 'Score Risque',
        'Score Fraude', 'Niveau Risque', 'Alerte Fraude',
        'Recommandation', 'Date Soumission', 'Conseiller'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Données
    for row_idx, dossier in enumerate(queryset, 2):
        data = [
            dossier.reference,
            str(dossier.client),
            dossier.client.get_type_client_display(),
            float(dossier.montant_demande),
            dossier.duree_mois,
            dossier.get_objet_pret_display(),
            dossier.get_etat_display(),
            float(dossier.score_risque) if dossier.score_risque else None,
            float(dossier.score_fraude) if dossier.score_fraude else None,
            dossier.niveau_risque,
            'OUI' if dossier.alerte_fraude else 'NON',
            dossier.recommandation,
            dossier.date_soumission.strftime('%d/%m/%Y %H:%M') if dossier.date_soumission else '',
            str(dossier.conseiller) if dossier.conseiller else '',
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            if col == 4:
                cell.number_format = '#,##0'

    # Largeurs de colonnes
    widths = [15, 25, 15, 18, 12, 18, 15, 13, 13, 15, 13, 40, 18, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def get_dashboard_data(user=None):
    """Récupère les données agrégées pour le dashboard, filtrées par rôle."""
    from django.db.models import Count, Avg, Sum, Q
    from .models import get_user_role

    dossiers = DossierPret.objects.all()
    if user:
        role = get_user_role(user)
        if role in ('manager_risque', 'admin'):
            pass  # tout voir
        elif role == 'gestionnaire':
            dossiers = dossiers.filter(client__cree_par=user)
        else:  # conseiller
            dossiers = dossiers.filter(conseiller=user)

    total_dossiers = dossiers.count()
    dossiers_par_etat = dict(
        dossiers.values_list('etat').annotate(count=Count('id')).values_list('etat', 'count')
    )

    risque_distribution = dict(
        dossiers.exclude(niveau_risque='').values_list('niveau_risque')
        .annotate(count=Count('id')).values_list('niveau_risque', 'count')
    )

    score_moyen = dossiers.exclude(score_risque=None).aggregate(avg=Avg('score_risque'))['avg'] or 0
    montant_total = dossiers.aggregate(total=Sum('montant_demande'))['total'] or 0
    alertes_fraude = dossiers.filter(alerte_fraude=True).count()
    dossiers_recents = dossiers.select_related('client', 'conseiller')[:10]

    par_objet = dict(
        dossiers.values_list('objet_pret').annotate(count=Count('id'))
        .values_list('objet_pret', 'count')
    )

    # Stats de performance
    total_valides = dossiers.filter(etat='valide').count()
    total_refuses = dossiers.filter(etat='refuse').count()
    taux_approbation = round(total_valides / total_dossiers * 100, 1) if total_dossiers > 0 else 0

    return {
        'total_dossiers': total_dossiers,
        'dossiers_par_etat': dossiers_par_etat,
        'risque_distribution': risque_distribution,
        'score_moyen': round(float(score_moyen), 1),
        'montant_total': montant_total,
        'alertes_fraude': alertes_fraude,
        'dossiers_recents': dossiers_recents,
        'par_objet': par_objet,
        'total_valides': total_valides,
        'total_refuses': total_refuses,
        'taux_approbation': taux_approbation,
    }
